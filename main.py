import IPy
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from pyvirtualdisplay import Display


def auto_screen(http, ip_port):

    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-dev-shm-usage')
    
    browser = webdriver.Chrome(options=options)
    browser.get(http + "://" + ip_port)
    browser.get_screenshot_as_file("./web_screen/" + http + '_' + ip_port + ".png")
    title = browser.title
    browser.quit()

    return title

wb = load_workbook('target.xlsx', data_only=True)
ws = wb.active
sheet_ranges = wb[wb.sheetnames[0]]

id = 1
for col in sheet_ranges['A']:
    
    if(id > 1):
        if IPy.IP(sheet_ranges['B' + str(id)].value).version() == 4:
            sheet_ranges['D'+str(id)] = "IPV4"
            ip_port = str(sheet_ranges['B' + str(id)].value) + ":" + str(sheet_ranges['C' + str(id)].value)
        elif IPy.IP(sheet_ranges['B' + str(id)].value).version() == 6:
            sheet_ranges['D'+str(id)] = "IPV6"
            ip_port = "[" + str(sheet_ranges['B' + str(id)].value) + "]:" + str(sheet_ranges['C' + str(id)].value)
        else:
            ip_port = "Error"
        print(ip_port)
        if ip_port != "Error":
            try:
                sheet_ranges['F'+str(id)] = auto_screen("http", ip_port)
                sheet_ranges['E'+str(id)] = "http://" + ip_port
            except:
                sheet_ranges['F'+str(id)] = "WEB_ERROR"
            try:
                sheet_ranges['H'+str(id)] = auto_screen("https", ip_port)
                sheet_ranges['G'+str(id)] = "https://" + ip_port
            except:
                sheet_ranges['H'+str(id)] = "WEB_ERROR"
    id = id + 1
wb.save('target.xlsx')